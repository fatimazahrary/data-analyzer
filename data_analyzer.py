import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import os
from datetime import datetime

# ─── Configuration ─────────────────────────────────────────────────────────────

os.makedirs("output/charts", exist_ok=True)
os.makedirs("output/reports", exist_ok=True)

# ─── Données simulées (ventes e-commerce) ──────────────────────────────────────

data = {
    "Date": pd.date_range(start="2024-01-01", periods=12, freq="ME"),
    "Month": ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],
    "Revenue":    [42000, 38500, 51000, 47000, 63000, 58000, 71000, 69000, 55000, 61000, 78000, 95000],
    "Orders":     [320,   290,   410,   380,   510,   470,   580,   560,   440,   490,   630,   780],
    "Customers":  [210,   185,   270,   245,   340,   310,   390,   375,   295,   330,   420,   510],
    "Returns":    [12,    18,    15,    20,    22,    19,    25,    28,    16,    21,    30,    35],
    "Category":   ["Electronics","Clothing","Electronics","Home","Clothing","Electronics",
                   "Home","Clothing","Electronics","Home","Clothing","Electronics"],
}

df = pd.DataFrame(data)
df["Avg_Order_Value"] = (df["Revenue"] / df["Orders"]).round(2)
df["Return_Rate_%"]   = ((df["Returns"] / df["Orders"]) * 100).round(2)
df["New_Customers"]   = (df["Customers"] * 0.6).astype(int)
df["Repeat_Customers"]= df["Customers"] - df["New_Customers"]


# ─── KPIs ──────────────────────────────────────────────────────────────────────

def compute_kpis(df):
    kpis = {
        "Total Revenue":        f"${df['Revenue'].sum():,.0f}",
        "Total Orders":         f"{df['Orders'].sum():,}",
        "Total Customers":      f"{df['Customers'].sum():,}",
        "Avg Monthly Revenue":  f"${df['Revenue'].mean():,.0f}",
        "Avg Order Value":      f"${df['Avg_Order_Value'].mean():.2f}",
        "Avg Return Rate":      f"{df['Return_Rate_%'].mean():.1f}%",
        "Best Month":           df.loc[df['Revenue'].idxmax(), 'Month'],
        "Worst Month":          df.loc[df['Revenue'].idxmin(), 'Month'],
    }
    print("\n" + "=" * 45)
    print("         📊 KEY PERFORMANCE INDICATORS")
    print("=" * 45)
    for k, v in kpis.items():
        print(f"  {k:<25} {v:>15}")
    print("=" * 45)
    return kpis


# ─── Graphiques ────────────────────────────────────────────────────────────────

COLORS = ["#2980B9", "#2ECC71", "#E74C3C", "#E67E22", "#9B59B6"]

def plot_revenue_trend(df):
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.plot(df["Month"], df["Revenue"], marker="o", color=COLORS[0], linewidth=2.5, markersize=8)
    ax.fill_between(df["Month"], df["Revenue"], alpha=0.15, color=COLORS[0])
    for i, (m, r) in enumerate(zip(df["Month"], df["Revenue"])):
        ax.annotate(f"${r/1000:.0f}K", (m, r), textcoords="offset points",
                    xytext=(0, 10), ha="center", fontsize=9, color="#2C3E50")
    ax.set_title("Monthly Revenue Trend — 2024", fontsize=14, fontweight="bold", color="#2C3E50")
    ax.set_xlabel("Month"); ax.set_ylabel("Revenue ($)")
    ax.grid(axis="y", linestyle="--", alpha=0.5)
    ax.spines[["top","right"]].set_visible(False)
    plt.tight_layout()
    path = "output/charts/revenue_trend.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✅ Chart saved: {path}")

def plot_orders_vs_customers(df):
    fig, ax1 = plt.subplots(figsize=(12, 5))
    ax2 = ax1.twinx()
    bars = ax1.bar(df["Month"], df["Orders"], color=COLORS[0], alpha=0.7, label="Orders")
    line, = ax2.plot(df["Month"], df["Customers"], marker="s", color=COLORS[2],
                     linewidth=2.5, markersize=7, label="Customers")
    ax1.set_title("Orders vs Customers — 2024", fontsize=14, fontweight="bold", color="#2C3E50")
    ax1.set_ylabel("Orders", color=COLORS[0]); ax2.set_ylabel("Customers", color=COLORS[2])
    ax1.grid(axis="y", linestyle="--", alpha=0.4)
    ax1.spines[["top","right"]].set_visible(False)
    fig.legend(loc="upper left", bbox_to_anchor=(0.1, 0.95))
    plt.tight_layout()
    path = "output/charts/orders_vs_customers.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✅ Chart saved: {path}")

def plot_category_revenue(df):
    cat_revenue = df.groupby("Category")["Revenue"].sum().sort_values(ascending=False)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))
    bars = ax1.bar(cat_revenue.index, cat_revenue.values, color=COLORS[:len(cat_revenue)], edgecolor="white")
    for bar, val in zip(bars, cat_revenue.values):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 500,
                 f"${val/1000:.0f}K", ha="center", fontsize=10, fontweight="bold")
    ax1.set_title("Revenue by Category", fontsize=13, fontweight="bold", color="#2C3E50")
    ax1.set_ylabel("Revenue ($)"); ax1.grid(axis="y", linestyle="--", alpha=0.4)
    ax1.spines[["top","right"]].set_visible(False)
    wedges, texts, autotexts = ax2.pie(cat_revenue.values, labels=cat_revenue.index,
                                        autopct="%1.1f%%", colors=COLORS[:len(cat_revenue)],
                                        startangle=90, pctdistance=0.8)
    for t in autotexts: t.set_fontsize(10); t.set_fontweight("bold")
    ax2.set_title("Revenue Share by Category", fontsize=13, fontweight="bold", color="#2C3E50")
    plt.tight_layout()
    path = "output/charts/category_revenue.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✅ Chart saved: {path}")

def plot_return_rate(df):
    fig, ax = plt.subplots(figsize=(12, 4))
    colors = [COLORS[2] if r > 5 else COLORS[1] for r in df["Return_Rate_%"]]
    bars = ax.bar(df["Month"], df["Return_Rate_%"], color=colors, edgecolor="white")
    ax.axhline(y=5, color="orange", linestyle="--", linewidth=1.5, label="5% threshold")
    for bar, val in zip(bars, df["Return_Rate_%"]):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                f"{val}%", ha="center", fontsize=9)
    ax.set_title("Monthly Return Rate — 2024", fontsize=14, fontweight="bold", color="#2C3E50")
    ax.set_ylabel("Return Rate (%)"); ax.legend()
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.spines[["top","right"]].set_visible(False)
    plt.tight_layout()
    path = "output/charts/return_rate.png"
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✅ Chart saved: {path}")


# ─── Export Excel ──────────────────────────────────────────────────────────────

def export_to_excel(df, kpis):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "📊 Dashboard"

    blue  = PatternFill("solid", fgColor="2980B9")
    green = PatternFill("solid", fgColor="2ECC71")
    light = PatternFill("solid", fgColor="EBF5FB")
    bold_white = Font(bold=True, color="FFFFFF", size=11)
    bold_dark  = Font(bold=True, color="2C3E50", size=11)
    center = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="BDC3C7")
    bdr  = Border(left=side, right=side, top=side, bottom=side)

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "📊 E-COMMERCE PERFORMANCE DASHBOARD — 2024"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = blue
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Period: Jan–Dec 2024"
    ws["A2"].font = Font(italic=True, color="7F8C8D", size=10)
    ws["A2"].alignment = center

    # KPIs
    ws.append([])
    ws.append(["KPI", "Value"])
    kpi_header_row = ws.max_row
    for col in [1, 2]:
        c = ws.cell(row=kpi_header_row, column=col)
        c.fill = blue; c.font = bold_white; c.alignment = center; c.border = bdr

    for k, v in kpis.items():
        ws.append([k, v])
        row = ws.max_row
        ws.cell(row=row, column=1).font = Font(bold=True, color="2C3E50")
        ws.cell(row=row, column=2).fill = light
        ws.cell(row=row, column=2).font = Font(bold=True, color="2980B9")
        for col in [1, 2]:
            ws.cell(row=row, column=col).border = bdr
            ws.cell(row=row, column=col).alignment = center

    # Data table
    ws.append([])
    cols = ["Month","Revenue","Orders","Customers","Avg_Order_Value","Return_Rate_%"]
    ws.append(cols)
    header_row = ws.max_row
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=header_row, column=i)
        c.fill = green; c.font = bold_white; c.alignment = center; c.border = bdr

    for _, row_data in df[cols].iterrows():
        ws.append(list(row_data))
        r = ws.max_row
        for i in range(1, len(cols)+1):
            ws.cell(row=r, column=i).border = bdr
            ws.cell(row=r, column=i).alignment = center

    for i, w in enumerate([14,12,10,12,18,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"output/reports/dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"  ✅ Excel report saved: {filename}")


# ─── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n🚀 Running E-Commerce Data Analyzer...\n")
    kpis = compute_kpis(df)
    print("\n📈 Generating charts...")
    plot_revenue_trend(df)
    plot_orders_vs_customers(df)
    plot_category_revenue(df)
    plot_return_rate(df)
    print("\n📁 Exporting Excel report...")
    export_to_excel(df, kpis)
    print("\n🎉 All done! Check the output/ folder.\n")
